from flask import Flask, render_template, request, jsonify, send_file
import requests
import re
import json
import io
import csv
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill
from urllib.parse import unquote, urlparse
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ── Scrapling integration (browser-fingerprint fetcher) ───────────────────────
sys.path.insert(0, r'C:\Users\steli\Documents\Claude_Stuff\Add-Ons\Scrapling')
try:
    import warnings as _warnings
    with _warnings.catch_warnings():
        _warnings.simplefilter("ignore")
        from scrapling import Fetcher as _ScraplingFetcher
        _fetcher = _ScraplingFetcher()
    SCRAPLING_OK = True
except Exception as _scrapling_err:
    SCRAPLING_OK = False
    _fetcher = None

app = Flask(__name__)

BUSINESS_TYPES = {
    "restaurant":  {"tags": [("amenity", "restaurant")], "label": "Restaurant"},
    "cafe":        {"tags": [("amenity", "cafe")], "label": "Cafe"},
    "bar":         {"tags": [("amenity", "bar"), ("amenity", "pub")], "label": "Bar / Pub"},
    "fast_food":   {"tags": [("amenity", "fast_food")], "label": "Fast Food"},
    "sushi":       {"tags": [("amenity", "restaurant"), ("amenity", "fast_food")], "cuisine": "sushi", "label": "Sushi"},
    "pizza":       {"tags": [("amenity", "restaurant"), ("amenity", "fast_food")], "cuisine": "pizza", "label": "Pizza"},
    "burger":      {"tags": [("amenity", "fast_food")], "cuisine": "burger", "label": "Burger"},
    "bakery":      {"tags": [("shop", "bakery")], "label": "Bakery"},
    "gym":          {"tags": [("leisure", "fitness_centre"), ("leisure", "sports_centre"), ("amenity", "gym")], "label": "Gym / Fitness"},
    "pharmacy":     {"tags": [("amenity", "pharmacy"), ("healthcare", "pharmacy")], "label": "Pharmacy"},
    "dentist":      {"tags": [("amenity", "dentist"), ("healthcare", "dentist")], "label": "Dentist"},
    "hospital":     {"tags": [("amenity", "hospital"), ("amenity", "clinic"), ("healthcare", "hospital"), ("healthcare", "clinic")], "label": "Hospital / Clinic"},
    "beauty":       {"tags": [("shop", "beauty"), ("shop", "cosmetics")], "label": "Beauty Salon"},
    "hairdresser":  {"tags": [("shop", "hairdresser")], "label": "Hairdresser"},
    "barber":       {"tags": [("shop", "barber")], "label": "Barber"},
    "supermarket":  {"tags": [("shop", "supermarket"), ("shop", "convenience")], "label": "Supermarket"},
    "clothes":      {"tags": [("shop", "clothes"), ("shop", "fashion")], "label": "Clothes Shop"},
    "electronics":  {"tags": [("shop", "electronics"), ("shop", "computer"), ("shop", "mobile_phone")], "label": "Electronics"},
    "bank":         {"tags": [("amenity", "bank")], "label": "Bank"},
    "hotel":        {"tags": [("tourism", "hotel"), ("tourism", "hostel"), ("tourism", "guest_house"), ("tourism", "motel")], "label": "Hotel / Hostel"},
    "fuel":         {"tags": [("amenity", "fuel")], "label": "Gas Station"},
    "school":       {"tags": [("amenity", "school"), ("amenity", "college"), ("amenity", "university")], "label": "School"},
    "cinema":       {"tags": [("amenity", "cinema")], "label": "Cinema"},
    "nightclub":    {"tags": [("amenity", "nightclub"), ("amenity", "music_venue")], "label": "Nightclub"},
    "atm":          {"tags": [("amenity", "atm"), ("atm", "yes")], "label": "ATM"},
    "laundry":      {"tags": [("shop", "laundry"), ("shop", "dry_cleaning")], "label": "Laundry"},
    "books":        {"tags": [("shop", "books")], "label": "Bookshop"},
    "florist":      {"tags": [("shop", "florist")], "label": "Florist"},
    "optician":     {"tags": [("shop", "optician")], "label": "Optician"},
    "veterinary":   {"tags": [("amenity", "veterinary")], "label": "Veterinary"},
    "car_repair":   {"tags": [("shop", "car_repair")], "label": "Car Repair"},
    "jewelry":      {"tags": [("shop", "jewelry")], "label": "Jewelry"},
    "pet_shop":     {"tags": [("shop", "pet"), ("shop", "pet_supplies"), ("shop", "pets"), ("shop", "pet_shop")], "label": "Pet Shop"},
    "car_wash":     {"tags": [("amenity", "car_wash")], "label": "Car Wash"},
    "post_office":  {"tags": [("amenity", "post_office")], "label": "Post Office"},
    "parking":      {"tags": [("amenity", "parking"), ("amenity", "parking_space")], "label": "Parking"},
    "ice_cream":    {"tags": [("amenity", "ice_cream"), ("shop", "ice_cream")], "label": "Ice Cream"},
    "hardware":     {"tags": [("shop", "hardware"), ("shop", "doityourself")], "label": "Hardware Store"},
    "bicycle":      {"tags": [("shop", "bicycle")], "label": "Bicycle Shop"},
    "alcohol":      {"tags": [("shop", "wine"), ("shop", "alcohol"), ("shop", "beverages")], "label": "Wine / Alcohol"},
    "travel":       {"tags": [("shop", "travel_agency"), ("tourism", "information")], "label": "Travel Agency"},
    "massage":      {"tags": [("shop", "massage"), ("amenity", "spa")], "label": "Massage / Spa"},
    "toys":         {"tags": [("shop", "toys")], "label": "Toy Shop"},
    "sports_shop":  {"tags": [("shop", "sports"), ("shop", "outdoor")], "label": "Sports Shop"},
    "tattoo":       {"tags": [("shop", "tattoo"), ("shop", "piercing")], "label": "Tattoo / Piercing"},
    "photography":  {"tags": [("shop", "photo"), ("shop", "photography")], "label": "Photography"},
    "music_shop":   {"tags": [("shop", "musical_instrument"), ("shop", "music")], "label": "Music Shop"},
    "swimming":     {"tags": [("leisure", "swimming_pool"), ("sport", "swimming"), ("amenity", "public_bath")], "label": "Swimming Pool"},
    "park":         {"tags": [("leisure", "park"), ("leisure", "garden"), ("leisure", "nature_reserve")], "label": "Park / Garden"},
    "playground":   {"tags": [("leisure", "playground")], "label": "Playground"},
}

FIELD_LABELS = {
    "name":          "Business Name",
    "address":       "Address",
    "phone":         "Phone",
    "website":       "Website",
    "email":         "Email",
    "opening_hours": "Opening Hours",
    "rating":        "Rating",
    "cuisine":       "Cuisine",
    "latitude":      "Latitude",
    "longitude":     "Longitude",
}


def build_address(tags):
    parts = []
    street = tags.get("addr:street", "")
    housenumber = tags.get("addr:housenumber", "")
    if street:
        parts.append(f"{street} {housenumber}".strip())
    if tags.get("addr:city"):
        parts.append(tags["addr:city"])
    if tags.get("addr:postcode"):
        parts.append(tags["addr:postcode"])
    return ", ".join(parts)


def parse_element(element):
    tags = element.get("tags", {})
    lat = element.get("lat") or element.get("center", {}).get("lat", "")
    lon = element.get("lon") or element.get("center", {}).get("lon", "")
    return {
        "name":          tags.get("name", ""),
        "address":       build_address(tags),
        "phone":         tags.get("phone", tags.get("contact:phone", "")),
        "website":       tags.get("website", tags.get("contact:website", "")),
        "email":         tags.get("email", tags.get("contact:email", "")),
        "opening_hours": tags.get("opening_hours", ""),
        "rating":        tags.get("stars", tags.get("rating", "")),
        "cuisine":       tags.get("cuisine", ""),
        "latitude":      lat,
        "longitude":     lon,
    }


def search_nominatim_pois(label, area_type, area_data):
    """Secondary POI search via Nominatim — catches places Overpass may miss."""
    if area_type == "bbox":
        s, n, w, e = area_data
    else:
        lats = [pt[0] for pt in area_data]
        lons = [pt[1] for pt in area_data]
        s, n, w, e = min(lats), max(lats), min(lons), max(lons)

    try:
        resp = requests.get(
            "https://nominatim.openstreetmap.org/search",
            params={
                "q": label,
                "format": "jsonv2",
                "limit": 50,
                "bounded": 1,
                "viewbox": f"{w},{n},{e},{s}",  # west,north,east,south
                "addressdetails": 1,
                "extratags": 1,
            },
            headers={"User-Agent": "Scraper/1.0 (business-data-tool)"},
            timeout=15,
        )
        resp.raise_for_status()
    except requests.RequestException:
        return []

    results = []
    for r in resp.json():
        tags = r.get("extratags") or {}
        addr = r.get("address") or {}

        parts = []
        road = addr.get("road", "")
        house = addr.get("house_number", "")
        if road:
            parts.append(f"{road} {house}".strip())
        city = addr.get("city") or addr.get("town") or addr.get("village") or ""
        if city:
            parts.append(city)
        if addr.get("postcode"):
            parts.append(addr["postcode"])

        name = r.get("name") or r.get("display_name", "").split(",")[0].strip()

        results.append({
            "name":          name,
            "address":       ", ".join(parts),
            "phone":         tags.get("phone", tags.get("contact:phone", "")),
            "website":       tags.get("website", tags.get("contact:website", "")),
            "email":         tags.get("email", tags.get("contact:email", "")),
            "opening_hours": tags.get("opening_hours", ""),
            "rating":        tags.get("stars", tags.get("rating", "")),
            "cuisine":       tags.get("cuisine", ""),
            "latitude":      float(r["lat"]),
            "longitude":     float(r["lon"]),
        })
    return results


@app.route("/")
def index():
    return render_template("index.html", business_types=BUSINESS_TYPES)


@app.route("/api/geocode", methods=["POST"])
def geocode():
    city = request.json.get("city", "").strip()
    if not city:
        return jsonify({"error": "No location provided"}), 400

    try:
        resp = requests.get(
            "https://nominatim.openstreetmap.org/search",
            params={"q": city, "format": "json", "limit": 5},
            headers={"User-Agent": "Scraper/1.0 (business-data-tool)"},
            timeout=10,
        )
        resp.raise_for_status()
    except requests.RequestException as e:
        return jsonify({"error": f"Geocoding failed: {e}"}), 502

    results = resp.json()
    if not results:
        return jsonify({"error": "Location not found"}), 404

    out = []
    for r in results:
        out.append({
            "display_name": r["display_name"],
            "lat": float(r["lat"]),
            "lon": float(r["lon"]),
            "bbox": [float(x) for x in r["boundingbox"]],  # [south, north, west, east]
        })
    return jsonify({"results": out})


@app.route("/api/search", methods=["POST"])
def search():
    data = request.json
    business_key = data.get("business_type", "restaurant")
    area_type = data.get("area_type")   # "polygon" or "bbox"
    area_data = data.get("area_data")

    if not area_type or not area_data:
        return jsonify({"error": "No area selected"}), 400

    # Build area filter string
    if area_type == "polygon":
        poly_str = " ".join([f"{pt[0]} {pt[1]}" for pt in area_data])
        area_filter = f'(poly:"{poly_str}")'
    else:
        s, n, w, e = area_data
        area_filter = f"({s},{w},{n},{e})"

    # Build query parts (node + way for each tag pair)
    parts = []

    if business_key == "other":
        custom_value = data.get("custom_value", "").strip()
        if not custom_value:
            return jsonify({"error": "Please enter a search term"}), 400
        if "=" in custom_value:
            tag_key, tag_val = custom_value.split("=", 1)
            tag_pairs = [(tag_key.strip(), tag_val.strip())]
        else:
            val = custom_value.replace(" ", "_").lower()
            tag_pairs = [("amenity", val), ("shop", val), ("leisure", val), ("tourism", val)]
        for tag_key, tag_val in tag_pairs:
            parts.append(f'node["{tag_key}"="{tag_val}"]{area_filter};')
            parts.append(f'way["{tag_key}"="{tag_val}"]{area_filter};')
    else:
        bt = BUSINESS_TYPES.get(business_key)
        if not bt:
            return jsonify({"error": "Unknown business type"}), 400
        cuisine = bt.get("cuisine")
        for tag_key, tag_val in bt["tags"]:
            cuisine_filter = f'["cuisine"="{cuisine}"]' if cuisine else ""
            parts.append(f'node["{tag_key}"="{tag_val}"]{cuisine_filter}{area_filter};')
            parts.append(f'way["{tag_key}"="{tag_val}"]{cuisine_filter}{area_filter};')

    query = f"""[out:json][timeout:90][maxsize:536870912];
(
  {"".join(parts)}
);
out center;"""

    overpass_endpoints = [
        "https://overpass-api.de/api/interpreter",
        "https://overpass.kumi.systems/api/interpreter",
        "https://overpass.private.coffee/api/interpreter",
        "https://overpass.openstreetmap.ru/api/interpreter",
    ]

    resp = None
    last_error = None
    for endpoint in overpass_endpoints:
        try:
            resp = requests.post(endpoint, data=query, timeout=95)
            resp.raise_for_status()
            break
        except requests.RequestException as e:
            last_error = e
            continue

    if resp is None or not resp.ok:
        return jsonify({"error": "All map data servers are currently busy. Please wait a moment and try again."}), 502

    elements = resp.json().get("elements", [])

    # Dedup key: (lowercase name, lat rounded to ~10m, lon rounded to ~10m)
    def dedup_key(b):
        return (
            (b["name"] or "").lower().strip(),
            round(float(b["latitude"]  or 0), 4),
            round(float(b["longitude"] or 0), 4),
        )

    seen = set()
    businesses = []
    for el in elements:
        b = parse_element(el)
        k = dedup_key(b)
        if k not in seen:
            seen.add(k)
            businesses.append(b)

    # Secondary source: Nominatim POI search
    label = BUSINESS_TYPES[business_key]["label"] if business_key != "other" else data.get("custom_value", "")
    for b in search_nominatim_pois(label, area_type, area_data):
        k = dedup_key(b)
        if k not in seen:
            seen.add(k)
            businesses.append(b)

    return jsonify({"businesses": businesses, "count": len(businesses)})


@app.route("/api/export", methods=["POST"])
def export():
    data = request.json
    businesses = data.get("businesses", [])
    fields = data.get("fields", [])
    fmt = data.get("format", "csv")

    if not businesses:
        return jsonify({"error": "No data to export"}), 400
    if not fields:
        return jsonify({"error": "No fields selected"}), 400

    headers = [FIELD_LABELS.get(f, f) for f in fields]
    rows = [[str(b.get(f, "") or "") for f in fields] for b in businesses]

    if fmt == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(rows)
        return send_file(
            io.BytesIO(output.getvalue().encode("utf-8-sig")),
            mimetype="text/csv",
            as_attachment=True,
            download_name="businesses.csv",
        )

    elif fmt == "xlsx":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Businesses"

        header_fill = PatternFill("solid", fgColor="2563EB")
        header_font = Font(bold=True, color="FFFFFF")
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for row in rows:
            ws.append(row)

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="businesses.xlsx",
        )

    elif fmt == "txt":
        lines = ["\t".join(headers)]
        for row in rows:
            lines.append("\t".join(row))
        content = "\n".join(lines)
        return send_file(
            io.BytesIO(content.encode("utf-8")),
            mimetype="text/plain",
            as_attachment=True,
            download_name="businesses.txt",
        )

    return jsonify({"error": "Invalid format"}), 400


import unicodedata

BROWSER_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}

# ── Constants ─────────────────────────────────────────────
SKIP_WEBSITE_DOMAINS = [
    "google.com", "maps.google", "wikipedia.org", "wikidata.org",
    "openstreetmap.org", "duckduckgo.com", "bing.com", "yahoo.com",
    "booking.com", "expedia.com", "airbnb.com",
    "yelp.com", "tripadvisor.com", "foursquare.com", "zomato.com",
    "trustpilot.com", "yellowpages.com", "whitepages.com", "hotfrog.com",
    "facebook.com", "instagram.com", "twitter.com", "x.com",
    "linkedin.com", "youtube.com", "tiktok.com", "pinterest.com",
]

FAKE_EMAIL_WORDS = [
    "example", "domain", "yourname", "sentry", "schema", "wixpress",
    "squarespace", "wordpress", "cloudflare", "noreply", "no-reply",
    "webmaster", "postmaster", "mailer-daemon", "bounce",
    ".png", ".jpg", ".gif", ".svg", ".js", ".css",
]

PHONE_RE = re.compile(
    r"(?<!\d)(\+?[\d][\d\s.\-\(\)]{6,18}\d)(?!\d)"
)
EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")

# ── Helpers ───────────────────────────────────────────────

def clean_phones(raw_list):
    out = []
    for p in raw_list:
        p = p.strip()
        digits = re.sub(r"\D", "", p)
        if 7 <= len(digits) <= 15 and p not in out:
            out.append(p)
    return out


def clean_emails(raw_list):
    out = []
    for e in raw_list:
        if (len(e) < 80
                and not any(w in e.lower() for w in FAKE_EMAIL_WORDS)
                and e not in out):
            out.append(e)
    return out


def fetch_page(url, timeout=10):
    """
    Fetch URL using Scrapling Fetcher (browser fingerprinting, bot evasion).
    Returns the Scrapling Adaptor object on success, or None.
    Falls back to plain requests if Scrapling is unavailable.
    """
    if SCRAPLING_OK and _fetcher is not None:
        try:
            page = _fetcher.get(url, timeout=timeout, stealthy_headers=True)
            if page is not None:
                return page
        except Exception:
            pass
    # Fallback: wrap raw HTML in a fake object so callers get consistent interface
    return _fetch_html_raw(url, timeout)


def _fetch_html_raw(url, timeout=10):
    """Plain requests fallback — returns raw HTML string or None."""
    try:
        r = requests.get(url, headers=BROWSER_HEADERS, timeout=timeout,
                         verify=False, allow_redirects=True)
        if r.ok:
            return r.text[:150000]
    except Exception:
        pass
    return None


# Keep fetch_html as alias for backward compat with bing_rss / domain-guess callers
def fetch_html(url, timeout=10):
    result = fetch_page(url, timeout)
    if result is None:
        return None
    # If it's already a string (requests fallback), return as-is
    if isinstance(result, str):
        return result
    # Scrapling Response: get raw HTML
    return _page_to_html(result) or None


def extract_jsonld(html):
    """Extract phone/email from schema.org JSON-LD blocks (raw HTML string)."""
    phone = email = ""
    for raw in re.findall(
        r'<script[^>]+type=["\']application/ld\+json["\'][^>]*>(.*?)</script>',
        html, re.DOTALL | re.IGNORECASE
    ):
        try:
            obj = json.loads(raw.strip())
            if isinstance(obj, list):
                obj = obj[0] if obj else {}
            phone = phone or str(obj.get("telephone") or "").strip()
            email = email or str(obj.get("email") or "").strip()
        except Exception:
            pass
    return phone, email


def _extract_contact_from_html(html):
    """Extract email + phone from raw HTML string via regex."""
    phone_jld, email_jld = extract_jsonld(html)

    tel_hrefs    = re.findall(r'href=["\']tel:([^"\'>\s]+)["\']', html, re.IGNORECASE)
    mailto_hrefs = re.findall(r'href=["\']mailto:([^"\'?>\s]+)["\']', html, re.IGNORECASE)

    clean = re.sub(r"<(script|style)[^>]*>.*?</\1>", " ", html,
                   flags=re.DOTALL | re.IGNORECASE)
    for obf, rep in [("&#64;","@"),("%40","@"),("[at]","@"),("(at)","@"),(" AT ","@"),(" at ","@")]:
        clean = clean.replace(obf, rep)
    plain = re.sub(r"<[^>]+>", " ", clean)

    phones = clean_phones(tel_hrefs + PHONE_RE.findall(plain))
    emails = clean_emails(mailto_hrefs + EMAIL_RE.findall(plain))

    return (
        email_jld or (emails[0] if emails else ""),
        phone_jld or (phones[0] if phones else ""),
    )


def _extract_contact_with_scrapling(page):
    """
    Extract email + phone from a Scrapling Adaptor using CSS selectors.
    Much more precise than regex: pulls actual tel:/mailto: hrefs directly.
    """
    phone_jld = email_jld = ""

    # 1. JSON-LD via CSS selector
    try:
        for script in page.css('script[type="application/ld+json"]'):
            try:
                text = getattr(script, 'text', None) or getattr(script, 'get_all_text', lambda: "")()
                obj  = json.loads(text.strip())
                if isinstance(obj, list):
                    obj = obj[0] if obj else {}
                phone_jld = phone_jld or str(obj.get("telephone") or "").strip()
                email_jld = email_jld or str(obj.get("email") or "").strip()
            except Exception:
                pass
    except Exception:
        pass

    # 2. tel: links via CSS selector
    tel_hrefs = []
    try:
        for a in page.css('a[href^="tel:"]'):
            href = ""
            try:
                href = a.attrib.get('href', '') if hasattr(a, 'attrib') else a['href']
            except Exception:
                try:
                    href = str(a.get('href') or "")
                except Exception:
                    pass
            val = href.replace('tel:', '').strip()
            if val:
                tel_hrefs.append(val)
    except Exception:
        pass

    # 3. mailto: links via CSS selector
    mailto_hrefs = []
    try:
        for a in page.css('a[href^="mailto:"]'):
            href = ""
            try:
                href = a.attrib.get('href', '') if hasattr(a, 'attrib') else a['href']
            except Exception:
                try:
                    href = str(a.get('href') or "")
                except Exception:
                    pass
            val = href.replace('mailto:', '').split('?')[0].strip()
            if val:
                mailto_hrefs.append(val)
    except Exception:
        pass

    # 4. Full text for regex fallback
    try:
        text = page.get_all_text() or ""
    except Exception:
        text = ""

    phones = clean_phones(tel_hrefs + PHONE_RE.findall(text))
    emails = clean_emails(mailto_hrefs + EMAIL_RE.findall(text))

    return (
        email_jld or (emails[0] if emails else ""),
        phone_jld or (phones[0] if phones else ""),
    )


def extract_contact(page_or_html):
    """Extract email + phone from a Scrapling Adaptor or raw HTML string."""
    if isinstance(page_or_html, str):
        return _extract_contact_from_html(page_or_html)
    if page_or_html is None:
        return "", ""
    # Scrapling Adaptor
    if SCRAPLING_OK:
        try:
            return _extract_contact_with_scrapling(page_or_html)
        except Exception:
            pass
    # Fallback: get HTML string from adaptor
    html = _page_to_html(page_or_html)
    return _extract_contact_from_html(html) if html else ("", "")


def _page_to_html(page):
    """Convert a Scrapling Response to raw HTML string."""
    try:
        h = str(page.html_content or "")
        if h:
            return h[:150000]
    except Exception:
        pass
    try:
        h = page.body.decode(page.encoding or "utf-8", errors="replace")
        return h[:150000]
    except Exception:
        pass
    try:
        return str(page)[:150000]
    except Exception:
        return ""


def find_socials_in_html_str(html):
    """Return (facebook_url, instagram_url) found as links in raw HTML."""
    def norm(pattern):
        m = re.search(pattern, html, re.IGNORECASE)
        if not m:
            return ""
        u = m.group(1)
        return u if u.startswith("http") else "https://" + u

    fb = norm(r'href=["\']((https?://)?(?:www\.)?facebook\.com/(?!sharer|share|dialog|tr\?)[^"\'?\s]+)["\']')
    ig = norm(r'href=["\']((https?://)?(?:www\.)?instagram\.com/[^"\'?\s]+)["\']')
    return fb, ig


def find_socials(page_or_html):
    """Return (facebook_url, instagram_url) from Scrapling Adaptor or raw HTML."""
    fb = ig = ""
    if not isinstance(page_or_html, str) and page_or_html is not None and SCRAPLING_OK:
        # Use CSS selectors for precision
        try:
            for a in page_or_html.css('a[href*="facebook.com"]'):
                try:
                    href = a.attrib.get('href', '') if hasattr(a, 'attrib') else a['href']
                except Exception:
                    href = ""
                if href and "/sharer" not in href and "/tr?" not in href and "/dialog/" not in href:
                    fb = href if href.startswith("http") else "https://" + href
                    break
        except Exception:
            pass
        try:
            for a in page_or_html.css('a[href*="instagram.com"]'):
                try:
                    href = a.attrib.get('href', '') if hasattr(a, 'attrib') else a['href']
                except Exception:
                    href = ""
                if href:
                    ig = href if href.startswith("http") else "https://" + href
                    break
        except Exception:
            pass
        if fb or ig:
            return fb, ig
        # Fallback to HTML string
        html = _page_to_html(page_or_html)
        return find_socials_in_html_str(html)

    html = page_or_html if isinstance(page_or_html, str) else ""
    return find_socials_in_html_str(html)


# Keep old name as alias for any call sites that use it directly
find_socials_in_html = find_socials_in_html_str


def scrape_url(url):
    """Fetch a URL with Scrapling and return (email, phone, fb_url, ig_url)."""
    page = fetch_page(url)
    if page is None:
        return "", "", "", ""
    email, phone = extract_contact(page)
    fb, ig       = find_socials(page)
    return email, phone, fb, ig


NOM_HEADERS = {"User-Agent": "Scraper/1.0 (business-data-tool)"}


def nominatim_lookup(name, lat, lon):
    """
    Use Nominatim bounded search near the known coordinates.
    Returns dict of OSM extratags if found, else {}.
    """
    try:
        r = requests.get(
            "https://nominatim.openstreetmap.org/search",
            params={
                "q": name,
                "format": "jsonv2",
                "limit": 5,
                "viewbox": f"{float(lon)-0.02},{float(lat)+0.02},{float(lon)+0.02},{float(lat)-0.02}",
                "bounded": 1,
                "extratags": 1,
                "addressdetails": 0,
            },
            headers=NOM_HEADERS,
            timeout=10,
        )
        if r.ok:
            for res in r.json():
                tags = res.get("extratags") or {}
                if any(tags.get(k) for k in ["phone","website","email","contact:phone","contact:website"]):
                    return tags
            # Return first result's tags even if empty, for website field
            results = r.json()
            if results:
                return results[0].get("extratags") or {}
    except Exception:
        pass
    return {}


def overpass_lookup(name, lat, lon):
    """
    Query Overpass for a node/way named `name` within 150m of coordinates.
    Returns OSM tags dict or {}.
    """
    escaped = name.replace('"', '\\"')
    query = (
        f'[out:json][timeout:15];'
        f'('
        f'node["name"~"{escaped}",i](around:150,{lat},{lon});'
        f'way["name"~"{escaped}",i](around:150,{lat},{lon});'
        f');'
        f'out tags;'
    )
    endpoints = [
        "https://overpass-api.de/api/interpreter",
        "https://overpass.kumi.systems/api/interpreter",
        "https://overpass.private.coffee/api/interpreter",
    ]
    for ep in endpoints:
        try:
            r = requests.post(ep, data=query, timeout=18)
            if r.ok and r.text.strip():
                elements = r.json().get("elements", [])
                for el in elements:
                    tags = el.get("tags", {})
                    if any(tags.get(k) for k in ["phone","website","email","contact:phone","contact:website"]):
                        return tags
                if elements:
                    return elements[0].get("tags", {})
            break
        except Exception:
            continue
    return {}


def extract_from_osm_tags(tags):
    """Pull website, phone, email out of OSM tags dict."""
    website = (tags.get("website") or tags.get("contact:website") or
               tags.get("url") or "").strip()
    phone   = (tags.get("phone") or tags.get("contact:phone") or
               tags.get("telephone") or "").strip()
    email   = (tags.get("email") or tags.get("contact:email") or "").strip()
    return website, phone, email


def bing_rss(query, limit=8):
    """Bing RSS with forced en-US market to avoid geo-redirection."""
    try:
        r = requests.get(
            "https://www.bing.com/search",
            params={"q": query, "format": "rss", "setlang": "en-US",
                    "setmkt": "en-US", "cc": "US"},
            headers=BROWSER_HEADERS,
            timeout=12,
        )
        if not r.ok:
            return [], ""
        items   = re.findall(r'<item>(.*?)</item>', r.text, re.DOTALL)
        urls    = []
        snippets = []
        for item in items:
            lm = re.search(r'<link>(https?://[^<]+)</link>', item)
            dm = re.search(r'<description>(.*?)</description>', item, re.DOTALL)
            if lm:
                urls.append(lm.group(1).strip())
            if dm:
                snippets.append(re.sub(r'<[^>]+>', ' ', dm.group(1)))
        return urls[:limit], " ".join(snippets)
    except Exception:
        return [], ""


def guess_domain_urls(name, address=""):
    """Generate plausible website URLs from a business name + country TLD."""
    n = unicodedata.normalize("NFD", name.lower())
    n = "".join(c for c in n if unicodedata.category(c) != "Mn")
    n = re.sub(r"[^a-z0-9\s]", "", n).strip()
    slug      = re.sub(r"\s+", "", n)
    slug_dash = re.sub(r"\s+", "-", n)

    # Guess country TLD from address
    country_tlds = {
        "greece": ".gr", "athens": ".gr", "thessaloniki": ".gr",
        "germany": ".de", "berlin": ".de", "munich": ".de",
        "france": ".fr", "paris": ".fr",
        "italy": ".it", "rome": ".it", "milan": ".it",
        "spain": ".es", "madrid": ".es", "barcelona": ".es",
        "netherlands": ".nl", "amsterdam": ".nl",
        "belgium": ".be", "brussels": ".be",
        "uk": ".co.uk", "london": ".co.uk", "england": ".co.uk",
        "australia": ".com.au", "sydney": ".com.au",
        "portugal": ".pt", "lisbon": ".pt",
    }
    extra_tlds = []
    addr_lower = address.lower()
    for keyword, tld in country_tlds.items():
        if keyword in addr_lower:
            extra_tlds.append(tld)

    tlds = [".com", ".net", ".org", ".eu", ".io"] + extra_tlds
    out = []
    for t in tlds:
        out.append(f"https://www.{slug}{t}")
        if slug_dash != slug:
            out.append(f"https://www.{slug_dash}{t}")
        out.append(f"https://{slug}{t}")
    return out


def find_website_for_business(name, address):
    """Find official website via Bing RSS then domain guessing."""
    location = ", ".join(address.split(",")[:2]) if address else ""
    queries = [
        f'"{name}" {location}',
        f'"{name}" {location} official website',
        f'{name} {location} contact phone',
    ]

    all_snippets = ""
    for query in queries:
        urls, snips = bing_rss(query, limit=10)
        all_snippets += " " + snips
        good = [u for u in urls
                if not any(s in u.lower() for s in SKIP_WEBSITE_DOMAINS)]
        if good:
            return good[0], all_snippets

    # Domain guessing
    for url in guess_domain_urls(name, address):
        html = fetch_html(url, timeout=6)
        if html:
            return url, all_snippets

    return "", all_snippets


def scrape_directory(url):
    """Scrape a directory/aggregator page for phone + email only."""
    page = fetch_page(url)
    if page is None:
        return "", ""
    email, phone = extract_contact(page)
    return email, phone


def find_on_directories(name, address):
    """Search Bing for the business on TripAdvisor, Yelp, Foursquare, Facebook."""
    location = ", ".join(address.split(",")[:2]) if address else ""
    email = phone = fb_url = ig_url = ""

    dir_sites = [
        "tripadvisor.com", "yelp.com", "foursquare.com",
        "zomato.com", "happycow.net", "thefork.com",
    ]
    social_sites = ["facebook.com", "instagram.com"]

    # Search all directory + social sites for this business
    for site in dir_sites + social_sites:
        if email and phone:
            break
        urls, snips = bing_rss(f'site:{site} "{name}" {location}', limit=3)
        for url in urls:
            if site not in url:
                continue
            if "facebook.com" in url:
                if not fb_url and "/sharer" not in url and "/tr?" not in url:
                    fb_url = url
                    continue
            if "instagram.com" in url:
                if not ig_url:
                    ig_url = url
                    continue
            # For directory sites scrape for phone/email
            e, p = scrape_directory(url)
            email = email or e
            phone = phone or p
            # Also check snippets for phones
            if not phone:
                sp = clean_phones(PHONE_RE.findall(snips))
                phone = phone or (sp[0] if sp else "")

    # Scrape social pages
    for surl in filter(None, [fb_url, ig_url]):
        if email and phone:
            break
        e, p = scrape_directory(surl)
        email = email or e
        phone = phone or p

    return email, phone, fb_url, ig_url


@app.route("/api/enrich", methods=["POST"])
def enrich():
    data    = request.json or {}
    name    = data.get("name", "").strip()
    address = data.get("address", "").strip()
    lat     = data.get("lat", "")
    lon     = data.get("lon", "")
    current = data.get("current", {})

    if not name:
        return jsonify({}), 400

    result  = {"website": "", "phone": "", "email": ""}
    email   = current.get("email", "")
    phone   = current.get("phone", "")
    website = current.get("website", "")
    fb_url  = ig_url = ""
    snippets = ""

    # ── 1. Nominatim bounded search (uses known lat/lon) ──
    #    Most reliable: directly queries OSM data by coordinates
    if lat and lon and (not website or not phone or not email):
        nom_tags = nominatim_lookup(name, lat, lon)
        w, p, e  = extract_from_osm_tags(nom_tags)
        website  = website or w
        phone    = phone   or p
        email    = email   or e

    # ── 2. Overpass radius query (deeper OSM tag coverage) ─
    if lat and lon and (not website or not phone or not email):
        osm_tags = overpass_lookup(name, lat, lon)
        w, p, e  = extract_from_osm_tags(osm_tags)
        website  = website or w
        phone    = phone   or p
        email    = email   or e

    # Update result with what we found so far
    result["website"] = website

    # ── 3. Find website via search + domain guessing ──────
    if not website:
        website, snippets = find_website_for_business(name, address)
        result["website"] = website

    # ── 4. Scrape the website homepage ───────────────────
    if website and (not email or not phone):
        e, p, fb, ig = scrape_url(website)
        email  = email  or e
        phone  = phone  or p
        fb_url = fb_url or fb
        ig_url = ig_url or ig

    # ── 5. Scrape contact/about sub-pages ────────────────
    if website and (not email or not phone):
        parsed = urlparse(website.rstrip("/"))
        origin = f"{parsed.scheme}://{parsed.netloc}"
        for path in ["/contact", "/contact-us", "/about", "/about-us",
                     "/kontakt", "/info", "/impressum", "/reach-us"]:
            if email and phone:
                break
            e, p, fb2, ig2 = scrape_url(origin + path)
            email  = email  or e
            phone  = phone  or p
            fb_url = fb_url or fb2
            ig_url = ig_url or ig2

    # ── 6. Directories + social pages ────────────────────
    if not email or not phone:
        de, dp, fb2, ig2 = find_on_directories(name, address)
        email  = email  or de
        phone  = phone  or dp
        fb_url = fb_url or fb2
        ig_url = ig_url or ig2

    for surl in filter(None, [fb_url, ig_url]):
        if email and phone:
            break
        e, p, _, _ = scrape_url(surl)
        email = email or e
        phone = phone or p

    # ── 7. Phone from search snippets ────────────────────
    if not phone and snippets:
        sp = clean_phones(PHONE_RE.findall(snippets))
        phone = phone or (sp[0] if sp else "")

    if not current.get("email"):
        result["email"] = email
    if not current.get("phone"):
        result["phone"] = phone

    return jsonify(result)


if __name__ == "__main__":
    app.run(debug=True, port=5000)
